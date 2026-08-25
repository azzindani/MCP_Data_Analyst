"""data_ingest engine — spreadsheet ingestion logic. Zero MCP imports."""

from __future__ import annotations

import csv
import logging
import sys
from pathlib import Path

_ROOT = str(Path(__file__).resolve().parents[2])
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

import pandas as pd

from shared.file_utils import (
    atomic_write,
    atomic_write_text,
    embed_content,
    error_text,
    get_default_output_dir,
    hint_for_error,
    resolve_path,
)
from shared.platform_utils import get_max_results
from shared.progress import fail, info, ok, warn
from shared.receipt import append_receipt
from shared.version_control import snapshot

logger = logging.getLogger(__name__)
logging.basicConfig(stream=sys.stderr, level=logging.WARNING)

_XLSX_EXTS = {".xlsx", ".ods"}
_ALL_INPUT_EXTS = {".xlsx", ".ods", ".csv", ".json", ".parquet"}
_OUTPUT_FMTS = {"csv", "json", "parquet", "excel"}
_FMT_EXT = {"csv": ".csv", "json": ".json", "parquet": ".parquet", "excel": ".xlsx"}


def _token_estimate(obj: object) -> int:
    return len(str(obj)) // 4


def _widest_row(path: Path) -> int:
    """Field count of the widest row in a CSV, honouring quoting."""
    width = 0
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.reader(fh):
            width = max(width, len(row))
    return width


def _first_row(path: Path) -> list[str]:
    """The first row's fields, honouring quoting."""
    with open(path, newline="", encoding="utf-8", errors="replace") as fh:
        for row in csv.reader(fh):
            return row
    return []


def read_csv_ragged(path: Path, header: int | None = 0) -> pd.DataFrame:
    """Read a CSV whose rows disagree about how many fields they have.

    pandas fixes the column count from the first row it reads and raises
    ParserError the moment a later row is wider:

        Error tokenizing data. C error: Expected 1 fields in line 3, saw 16

    That is the ordinary shape of the files this server exists to repair -- a
    title line above a table, a sheet exported with a short header row, a
    "generated on ..." banner. So promote_header(), trim_empty() and
    normalize_headers() each refused the exact input they were written for,
    and the error named a pandas internal rather than anything the caller
    could act on.

    The well-formed path is unchanged and pays nothing: a plain read is tried
    first, and the file is only re-scanned for its true width after pandas has
    already refused it. Short rows are then padded with NaN, which is what a
    spreadsheet shows for the same cells.
    """
    try:
        return pd.read_csv(str(path), header=header)
    except pd.errors.ParserError:
        # Only a width the whole file agrees on can rescue this; if the file
        # has no rows at all, the original ParserError is the honest answer.
        width = _widest_row(path)
        if width == 0:
            raise

    if header is None:
        return pd.read_csv(str(path), names=range(width))

    # Take the header off with the csv module and let pandas infer dtypes from
    # the data rows alone -- leaving the header in place would make every
    # column object-typed.
    names = _first_row(path)
    columns = [
        str(names[i]).strip() if i < len(names) and str(names[i]).strip() else f"Unnamed: {i}" for i in range(width)
    ]
    frame = pd.read_csv(str(path), names=range(width), skiprows=1)
    frame.columns = columns
    return frame


def _resolve_sheet(wb, sheet: str):
    """Return (ws, sheet_name) from an openpyxl workbook. sheet may be name or int-as-str."""
    names = wb.sheetnames
    if not sheet:
        ws = wb.active
        return ws, ws.title
    if sheet.lstrip("-").isdigit():
        idx = int(sheet)
        if idx < 0 or idx >= len(names):
            return None, None
        return wb[names[idx]], names[idx]
    if sheet not in names:
        return None, None
    return wb[sheet], sheet


def _sheet_to_df(path: Path, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
    """Read a single sheet into a DataFrame regardless of file format."""
    ext = path.suffix.lower()
    if ext == ".ods":
        return pd.read_excel(str(path), sheet_name=sheet_name, header=header_row, engine="odf")
    return pd.read_excel(str(path), sheet_name=sheet_name, header=header_row, engine="openpyxl")


def _find_tables(ws, min_rows: int, min_cols: int) -> list[dict]:
    """Detect bounding boxes of separate tables in a worksheet."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return []

    # Sequential iter_rows(), not random ws.cell(row, col) access — read_only
    # worksheets stream lazily and random cell access can stall or crawl.
    occupied = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
        occupied.append([val is not None and str(val).strip() != "" for val in row])

    # Find contiguous non-empty row groups
    row_groups: list[tuple[int, int]] = []
    in_group = False
    group_start = 0
    for i, row in enumerate(occupied):
        if any(row):
            if not in_group:
                group_start = i
                in_group = True
        else:
            if in_group:
                row_groups.append((group_start, i - 1))
                in_group = False
    if in_group:
        row_groups.append((group_start, len(occupied) - 1))

    tables = []
    for rs, re in row_groups:
        if (re - rs + 1) < min_rows:
            continue
        col_occ = [False] * max_col
        for r in range(rs, re + 1):
            for c in range(max_col):
                if occupied[r][c]:
                    col_occ[c] = True
        in_col = False
        cs = 0
        for c, occ in enumerate(col_occ):
            if occ:
                if not in_col:
                    cs = c
                    in_col = True
            else:
                if in_col:
                    if (c - cs) >= min_cols:
                        tables.append(
                            {
                                "row_start": rs,
                                "row_end": re,
                                "col_start": cs,
                                "col_end": c - 1,
                                "rows": re - rs + 1,
                                "cols": c - cs,
                            }
                        )
                    in_col = False
        if in_col and (max_col - cs) >= min_cols:
            tables.append(
                {
                    "row_start": rs,
                    "row_end": re,
                    "col_start": cs,
                    "col_end": max_col - 1,
                    "rows": re - rs + 1,
                    "cols": max_col - cs,
                }
            )
    for i, t in enumerate(tables):
        t["index"] = i
    return tables


# ---------------------------------------------------------------------------
# 1. list_sheets
# ---------------------------------------------------------------------------


def list_sheets(file_path: str) -> dict:
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        ext = path.suffix.lower()
        if ext not in _XLSX_EXTS:
            return {
                "success": False,
                "error": f"Expected .xlsx or .ods, got {ext!r}",
                "hint": "Use convert_file() to convert CSV/JSON/Parquet to xlsx first.",
                "progress": [fail("Wrong file type", ext)],
                "token_estimate": 20,
            }

        if ext == ".ods":
            xl = pd.ExcelFile(str(path), engine="odf")
            sheet_names = xl.sheet_names
            sheets = []
            for name in sheet_names:
                df = xl.parse(name, header=None)
                sheets.append({"name": name, "rows": len(df), "cols": len(df.columns)})
        else:
            import openpyxl

            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            sheets = []
            for name in sheet_names:
                ws = wb[name]
                sheets.append({"name": name, "rows": ws.max_row or 0, "cols": ws.max_column or 0})
            wb.close()

        max_r = get_max_results()
        truncated = len(sheets) > max_r
        if truncated:
            progress.append(warn("Truncated", f"Showing {max_r} of {len(sheets)} sheets"))
            sheets = sheets[:max_r]

        progress.append(ok("Listed sheets", f"{len(sheet_names)} sheet(s) in {path.name}"))
        result = {
            "success": True,
            "op": "list_sheets",
            "file": path.name,
            "file_path": str(path),
            "sheet_count": len(sheet_names),
            "sheets": sheets,
            "truncated": truncated,
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("list_sheets error")
        return {
            "success": False,
            "error": error_text(exc),
            "hint": hint_for_error(exc, "Check that the file is a valid Excel or ODS file."),
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }


# ---------------------------------------------------------------------------
# 2. extract_sheet
# ---------------------------------------------------------------------------


def extract_sheet(
    file_path: str,
    sheet: str = "",
    output_path: str = "",
    header_row: int = 0,
    dry_run: bool = False,
    return_content: bool = False,
) -> dict:
    backup = None
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        ext = path.suffix.lower()
        if ext not in _XLSX_EXTS:
            return {
                "success": False,
                "error": f"Expected .xlsx or .ods, got {ext!r}",
                "hint": "Use list_sheets() first to inspect available sheets.",
                "progress": [fail("Wrong file type", ext)],
                "token_estimate": 20,
            }

        # Resolve sheet name
        if ext == ".ods":
            xl = pd.ExcelFile(str(path), engine="odf")
            available = xl.sheet_names
            sheet_name = (
                available[0] if not sheet else (available[int(sheet)] if sheet.lstrip("-").isdigit() else sheet)
            )
            if sheet_name not in available:
                return {
                    "success": False,
                    "error": f"Sheet {sheet!r} not found",
                    "hint": f"Available: {available}. Call list_sheets() to inspect.",
                    "progress": [fail("Sheet not found", sheet)],
                    "token_estimate": 20,
                }
        else:
            import openpyxl

            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            ws, sheet_name = _resolve_sheet(wb, sheet)
            wb.close()
            if ws is None:
                available = wb.sheetnames if hasattr(wb, "sheetnames") else []
                return {
                    "success": False,
                    "error": f"Sheet {sheet!r} not found",
                    "hint": f"Available: {available}. Call list_sheets() to inspect.",
                    "progress": [fail("Sheet not found", sheet)],
                    "token_estimate": 20,
                }

        df = _sheet_to_df(path, sheet_name, header_row)
        out_dir = Path(output_path).parent if output_path else get_default_output_dir(str(path))
        stem = f"{path.stem}_{sheet_name}" if sheet_name != path.stem else path.stem
        out = Path(output_path) if output_path else out_dir / f"{stem}.csv"

        if dry_run:
            progress.append(info("Dry run — no changes written", path.name))
            result = {
                "success": True,
                "dry_run": True,
                "op": "extract_sheet",
                "sheet": sheet_name,
                "would_change": {"output_path": str(out), "rows": len(df), "cols": len(df.columns)},
                "progress": progress,
            }
            result["token_estimate"] = _token_estimate(result)
            return result

        if out.exists():
            backup = snapshot(str(out))
            progress.append(info("Snapshot created", Path(backup).name))

        out.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_text(out, df.to_csv(index=False))
        append_receipt(
            str(path),
            tool="extract_sheet",
            args={"sheet": sheet_name, "header_row": header_row},
            result=f"extracted {len(df)} rows to {out.name}",
            backup=backup or "",
        )
        progress.append(ok("Extracted sheet", f"{len(df)} rows → {out.name}"))
        result = {
            "success": True,
            "op": "extract_sheet",
            "file": path.name,
            "sheet": sheet_name,
            "output_path": str(out),
            "rows": len(df),
            "cols": len(df.columns),
            "backup": backup,
            "progress": progress,
        }
        embed_content(result, out, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("extract_sheet error")
        return {
            "success": False,
            "error": error_text(exc),
            "backup": backup,
            "hint": hint_for_error(exc, "Use list_sheets() to verify sheet names."),
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }


# ---------------------------------------------------------------------------
# 3. extract_all_sheets
# ---------------------------------------------------------------------------


def extract_all_sheets(file_path: str, output_dir: str = "", dry_run: bool = False) -> dict:
    backup = None
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        ext = path.suffix.lower()
        if ext not in _XLSX_EXTS:
            return {
                "success": False,
                "error": f"Expected .xlsx or .ods, got {ext!r}",
                "hint": "Use convert_file() to convert to xlsx first.",
                "progress": [fail("Wrong file type", ext)],
                "token_estimate": 20,
            }

        out_dir = Path(output_dir) if output_dir else get_default_output_dir(str(path))

        if ext == ".ods":
            xl = pd.ExcelFile(str(path), engine="odf")
            sheet_names = xl.sheet_names
        else:
            import openpyxl

            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()

        if dry_run:
            would = [{"sheet": s, "output_path": str(out_dir / f"{path.stem}_{s}.csv")} for s in sheet_names]
            progress.append(info("Dry run — no changes written", path.name))
            result = {
                "success": True,
                "dry_run": True,
                "op": "extract_all_sheets",
                "would_change": would,
                "sheet_count": len(sheet_names),
                "progress": progress,
            }
            result["token_estimate"] = _token_estimate(result)
            return result

        out_dir.mkdir(parents=True, exist_ok=True)

        extracted = []
        for name in sheet_names:
            df = _sheet_to_df(path, name)
            out = out_dir / f"{path.stem}_{name}.csv"
            atomic_write_text(out, df.to_csv(index=False))
            extracted.append({"sheet": name, "output_path": str(out), "rows": len(df)})
            progress.append(ok(f"Extracted {name}", f"{len(df)} rows → {out.name}"))

        append_receipt(
            str(path),
            tool="extract_all_sheets",
            args={"output_dir": str(out_dir)},
            result=f"extracted {len(extracted)} sheets",
            backup=backup,
        )
        result = {
            "success": True,
            "op": "extract_all_sheets",
            "file": path.name,
            "output_dir": str(out_dir),
            "extracted": extracted,
            "sheet_count": len(extracted),
            "backup": backup,
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("extract_all_sheets error")
        return {
            "success": False,
            "error": error_text(exc),
            "backup": backup,
            "hint": hint_for_error(exc, "Use restore_version() on the input file if a snapshot was taken."),
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }


# ---------------------------------------------------------------------------
# 4. detect_tables
# ---------------------------------------------------------------------------


def detect_tables(file_path: str, sheet: str = "", min_rows: int = 2, min_cols: int = 2) -> dict:
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        ext = path.suffix.lower()
        if ext not in _XLSX_EXTS:
            return {
                "success": False,
                "error": f"Expected .xlsx or .ods, got {ext!r}",
                "hint": "Use extract_sheet() to get a CSV first, then call detect_tables.",
                "progress": [fail("Wrong file type", ext)],
                "token_estimate": 20,
            }

        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws, sheet_name = _resolve_sheet(wb, sheet)
        if ws is None:
            available = wb.sheetnames
            wb.close()
            return {
                "success": False,
                "error": f"Sheet {sheet!r} not found",
                "hint": f"Available: {available}. Call list_sheets() to inspect.",
                "progress": [fail("Sheet not found", sheet)],
                "token_estimate": 20,
            }

        tables = _find_tables(ws, min_rows, min_cols)
        wb.close()

        max_r = get_max_results()
        truncated = len(tables) > max_r
        if truncated:
            progress.append(warn("Truncated", f"Showing {max_r} of {len(tables)} tables"))
            tables = tables[:max_r]

        progress.append(ok("Detected tables", f"{len(tables)} table(s) in sheet {sheet_name!r}"))
        result = {
            "success": True,
            "op": "detect_tables",
            "file": path.name,
            "sheet": sheet_name,
            "table_count": len(tables),
            "tables": tables,
            "truncated": truncated,
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("detect_tables error")
        return {
            "success": False,
            "error": error_text(exc),
            "hint": hint_for_error(exc, "Check the file is a valid Excel file."),
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }


# ---------------------------------------------------------------------------
# 5. extract_table
# ---------------------------------------------------------------------------


def extract_table(
    file_path: str,
    table_index: int = 0,
    sheet: str = "",
    output_path: str = "",
    min_rows: int = 2,
    min_cols: int = 2,
    dry_run: bool = False,
    return_content: bool = False,
) -> dict:
    backup = None
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        ext = path.suffix.lower()
        if ext not in _XLSX_EXTS:
            return {
                "success": False,
                "error": f"Expected .xlsx or .ods, got {ext!r}",
                "hint": "Use extract_sheet() first to get a CSV.",
                "progress": [fail("Wrong file type", ext)],
                "token_estimate": 20,
            }

        import openpyxl

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws, sheet_name = _resolve_sheet(wb, sheet)
        if ws is None:
            available = wb.sheetnames
            wb.close()
            return {
                "success": False,
                "error": f"Sheet {sheet!r} not found",
                "hint": f"Available: {available}. Call list_sheets() to inspect.",
                "progress": [fail("Sheet not found", sheet)],
                "token_estimate": 20,
            }

        tables = _find_tables(ws, min_rows=min_rows, min_cols=min_cols)
        wb.close()

        if table_index < 0 or table_index >= len(tables):
            return {
                "success": False,
                "error": f"table_index {table_index} out of range (found {len(tables)} tables)",
                "hint": "Call detect_tables() with the same min_rows/min_cols first to see available table indices.",
                "progress": [fail("Table index out of range", str(table_index))],
                "token_estimate": 20,
            }

        t = tables[table_index]
        bbox = {
            "row_start": t["row_start"],
            "row_end": t["row_end"],
            "col_start": t["col_start"],
            "col_end": t["col_end"],
        }

        # Read the whole sheet then slice
        df_full = _sheet_to_df(path, sheet_name, header_row=t["row_start"])
        col_slice = slice(t["col_start"], t["col_end"] + 1)
        # header_row consumed the first row; remaining rows start after header
        data_rows = t["row_end"] - t["row_start"]
        df = df_full.iloc[:data_rows, col_slice].reset_index(drop=True)

        out_dir = get_default_output_dir(str(path))
        out = Path(output_path) if output_path else out_dir / f"{path.stem}_{sheet_name}_table{table_index}.csv"

        if dry_run:
            progress.append(info("Dry run — no changes written", path.name))
            result = {
                "success": True,
                "dry_run": True,
                "op": "extract_table",
                "table_index": table_index,
                "sheet": sheet_name,
                "bounding_box": bbox,
                "would_change": {"output_path": str(out), "rows": len(df), "cols": len(df.columns)},
                "progress": progress,
            }
            result["token_estimate"] = _token_estimate(result)
            return result

        if out.exists():
            backup = snapshot(str(out))
            progress.append(info("Snapshot created", Path(backup).name))

        out.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_text(out, df.to_csv(index=False))
        append_receipt(
            str(path),
            tool="extract_table",
            args={"table_index": table_index, "sheet": sheet_name},
            result=f"extracted {len(df)} rows to {out.name}",
            backup=backup or "",
        )
        progress.append(ok("Extracted table", f"{len(df)} rows × {len(df.columns)} cols → {out.name}"))
        result = {
            "success": True,
            "op": "extract_table",
            "file": path.name,
            "table_index": table_index,
            "sheet": sheet_name,
            "bounding_box": bbox,
            "output_path": str(out),
            "rows": len(df),
            "cols": len(df.columns),
            "backup": backup,
            "progress": progress,
        }
        embed_content(result, out, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("extract_table error")
        return {
            "success": False,
            "error": error_text(exc),
            "backup": backup,
            "hint": hint_for_error(exc, "Call detect_tables() to verify table indices."),
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }


# ---------------------------------------------------------------------------
# 6. normalize_headers
# ---------------------------------------------------------------------------


def normalize_headers(
    file_path: str,
    lowercase: bool = True,
    replace_spaces: bool = True,
    output_path: str = "",
    dry_run: bool = False,
) -> dict:
    import re

    backup = None
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        if path.suffix.lower() != ".csv":
            return {
                "success": False,
                "error": f"Expected .csv, got {path.suffix!r}",
                "hint": "Use extract_sheet() or convert_file() to produce a CSV first.",
                "progress": [fail("Wrong file type", path.suffix)],
                "token_estimate": 20,
            }

        df = read_csv_ragged(path)
        old_cols = list(df.columns)
        new_cols: list[str] = []
        assigned: set[str] = set()

        for col in old_cols:
            new = col.strip()
            if lowercase:
                new = new.lower()
            if replace_spaces:
                new = new.replace(" ", "_")
            new = re.sub(r"_+", "_", new).strip("_") or "col"
            candidate = new
            counter = 2
            while candidate in assigned:
                candidate = f"{new}_{counter}"
                counter += 1
            assigned.add(candidate)
            new_cols.append(candidate)

        changes = {old: nw for old, nw in zip(old_cols, new_cols) if old != nw}
        out = resolve_path(output_path) if output_path else path

        if dry_run:
            progress.append(info("Dry run — no changes written", path.name))
            result = {
                "success": True,
                "dry_run": True,
                "op": "normalize_headers",
                "would_change": changes,
                "would_write": str(out),
                "progress": progress,
            }
            result["token_estimate"] = _token_estimate(result)
            return result

        # Only the source needs saving from itself; a named destination leaves
        # the caller's file untouched, so there is nothing to snapshot.
        if out == path:
            backup = snapshot(str(path))
            progress.append(info("Snapshot created", Path(backup).name))
        df.columns = new_cols
        out.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_text(out, df.to_csv(index=False))
        append_receipt(
            str(path),
            tool="normalize_headers",
            args={"lowercase": lowercase, "replace_spaces": replace_spaces},
            result=f"renamed {len(changes)} headers → {out.name}",
            backup=backup or "",
        )
        progress.append(ok("Normalized headers", f"{len(changes)} renamed → {out.name}"))
        result = {
            "success": True,
            "op": "normalize_headers",
            "file": path.name,
            "output_path": str(out),
            "changes": changes,
            "renamed_count": len(changes),
            "deduped_count": sum(
                1 for o, n in zip(old_cols, new_cols) if o != n and n != o.strip().lower().replace(" ", "_")
            ),
            "backup": backup or "",
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("normalize_headers error")
        return {
            "success": False,
            "error": error_text(exc),
            "backup": backup,
            "hint": hint_for_error(exc, "Use inspect_dataset() to verify column names first."),
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }


# ---------------------------------------------------------------------------
# 7. trim_empty
# ---------------------------------------------------------------------------


def trim_empty(file_path: str, output_path: str = "", dry_run: bool = False) -> dict:
    backup = None
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        if path.suffix.lower() != ".csv":
            return {
                "success": False,
                "error": f"Expected .csv, got {path.suffix!r}",
                "hint": "Use extract_sheet() or convert_file() to produce a CSV first.",
                "progress": [fail("Wrong file type", path.suffix)],
                "token_estimate": 20,
            }

        df = read_csv_ragged(path)
        rows_before = len(df)
        cols_before = len(df.columns)

        df = df.replace(r"^\s*$", pd.NA, regex=True)
        df = df.dropna(axis=1, how="all")
        df = df.dropna(axis=0, how="all")
        df = df.reset_index(drop=True)
        df = df.fillna("")

        rows_after = len(df)
        cols_after = len(df.columns)
        rows_dropped = rows_before - rows_after
        cols_dropped = cols_before - cols_after

        out = resolve_path(output_path) if output_path else path

        if dry_run:
            progress.append(info("Dry run — no changes written", path.name))
            result = {
                "success": True,
                "dry_run": True,
                "op": "trim_empty",
                "would_change": {"rows_to_drop": rows_dropped, "cols_to_drop": cols_dropped},
                "would_write": str(out),
                "progress": progress,
            }
            result["token_estimate"] = _token_estimate(result)
            return result

        if out == path:
            backup = snapshot(str(path))
            progress.append(info("Snapshot created", Path(backup).name))
        out.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_text(out, df.to_csv(index=False))
        append_receipt(
            str(path),
            tool="trim_empty",
            args={},
            result=f"dropped {rows_dropped} rows, {cols_dropped} cols → {out.name}",
            backup=backup or "",
        )
        progress.append(ok("Trimmed empty", f"-{rows_dropped} rows, -{cols_dropped} cols → {out.name}"))
        result = {
            "success": True,
            "op": "trim_empty",
            "file": path.name,
            "output_path": str(out),
            "rows_before": rows_before,
            "rows_after": rows_after,
            "cols_before": cols_before,
            "cols_after": cols_after,
            "rows_dropped": rows_dropped,
            "cols_dropped": cols_dropped,
            "backup": backup or "",
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("trim_empty error")
        return {
            "success": False,
            "error": error_text(exc),
            "backup": backup,
            "hint": hint_for_error(exc, "Use inspect_dataset() to verify the file structure first."),
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }


# ---------------------------------------------------------------------------
# 8. promote_header
# ---------------------------------------------------------------------------


def promote_header(file_path: str, row_index: int = 0, output_path: str = "", dry_run: bool = False) -> dict:
    backup = None
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        if path.suffix.lower() != ".csv":
            return {
                "success": False,
                "error": f"Expected .csv, got {path.suffix!r}",
                "hint": "Use extract_sheet() or convert_file() to produce a CSV first.",
                "progress": [fail("Wrong file type", path.suffix)],
                "token_estimate": 20,
            }

        df = read_csv_ragged(path, header=None)
        if row_index < 0 or row_index >= len(df):
            return {
                "success": False,
                "error": f"row_index {row_index} out of range (file has {len(df)} rows)",
                "hint": f"Valid range: 0 to {len(df) - 1}.",
                "progress": [fail("Row index out of range", str(row_index))],
                "token_estimate": 20,
            }

        new_headers = [str(v) if pd.notna(v) else f"col_{i}" for i, v in enumerate(df.iloc[row_index])]
        rows_dropped = row_index + 1
        df = df.iloc[rows_dropped:].copy()
        df.columns = new_headers
        df = df.reset_index(drop=True)

        out = resolve_path(output_path) if output_path else path

        if dry_run:
            progress.append(info("Dry run — no changes written", path.name))
            result = {
                "success": True,
                "dry_run": True,
                "op": "promote_header",
                "would_change": {"new_headers": new_headers, "rows_dropped_above": rows_dropped},
                "would_write": str(out),
                "progress": progress,
            }
            result["token_estimate"] = _token_estimate(result)
            return result

        if out == path:
            backup = snapshot(str(path))
            progress.append(info("Snapshot created", Path(backup).name))
        out.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_text(out, df.to_csv(index=False))
        append_receipt(
            str(path),
            tool="promote_header",
            args={"row_index": row_index},
            result=f"promoted row {row_index} as header → {out.name}",
            backup=backup or "",
        )
        progress.append(ok("Promoted header", f"row {row_index} → columns; {rows_dropped} row(s) dropped"))
        result = {
            "success": True,
            "op": "promote_header",
            "file": path.name,
            "output_path": str(out),
            "promoted_row_index": row_index,
            "new_headers": new_headers,
            "rows_dropped_above": rows_dropped,
            "backup": backup or "",
            "progress": progress,
        }
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("promote_header error")
        return {
            "success": False,
            "error": error_text(exc),
            "backup": backup,
            "hint": hint_for_error(exc, "Use inspect_dataset() to verify row structure first."),
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }


# ---------------------------------------------------------------------------
# 9. flatten_merged_cells
# ---------------------------------------------------------------------------


def flatten_merged_cells(
    file_path: str,
    sheet: str = "",
    output_path: str = "",
    dry_run: bool = False,
    return_content: bool = False,
) -> dict:
    backup = None
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        # ODS does not have merged cells in the same sense; xlsx only
        if path.suffix.lower() != ".xlsx":
            return {
                "success": False,
                "error": f"Expected .xlsx, got {path.suffix!r}",
                "hint": "flatten_merged_cells only works on .xlsx files.",
                "progress": [fail("Wrong file type", path.suffix)],
                "token_estimate": 20,
            }

        import openpyxl

        # Must NOT use read_only=True — merged_cells requires full load
        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws, sheet_name = _resolve_sheet(wb, sheet)
        if ws is None:
            available = wb.sheetnames
            wb.close()
            return {
                "success": False,
                "error": f"Sheet {sheet!r} not found",
                "hint": f"Available: {available}. Call list_sheets() to inspect.",
                "progress": [fail("Sheet not found", sheet)],
                "token_estimate": 20,
            }

        merge_map: dict[tuple[int, int], object] = {}
        merged_count = len(list(ws.merged_cells.ranges))
        for merged_range in list(ws.merged_cells.ranges):
            top_val = ws.cell(merged_range.min_row, merged_range.min_col).value
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    merge_map[(r, c)] = top_val

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        rows_data = []
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                row_vals.append(merge_map.get((r, c), ws.cell(r, c).value))
            rows_data.append(row_vals)
        wb.close()

        if not rows_data:
            df = pd.DataFrame()
        else:
            headers = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(rows_data[0])]
            df = pd.DataFrame(rows_data[1:], columns=headers)

        out_dir = get_default_output_dir(str(path))
        out = Path(output_path) if output_path else out_dir / f"{path.stem}_{sheet_name}_flat.csv"

        if dry_run:
            progress.append(info("Dry run — no changes written", path.name))
            result = {
                "success": True,
                "dry_run": True,
                "op": "flatten_merged_cells",
                "sheet": sheet_name,
                "would_change": {
                    "merged_regions": merged_count,
                    "output_path": str(out),
                    "rows": len(df),
                    "cols": len(df.columns),
                },
                "progress": progress,
            }
            result["token_estimate"] = _token_estimate(result)
            return result

        if out.exists():
            backup = snapshot(str(out))
            progress.append(info("Snapshot created", Path(backup).name))

        out.parent.mkdir(parents=True, exist_ok=True)
        atomic_write_text(out, df.to_csv(index=False))
        append_receipt(
            str(path),
            tool="flatten_merged_cells",
            args={"sheet": sheet_name},
            result=f"flattened {merged_count} merged regions → {out.name}",
            backup=backup or "",
        )
        progress.append(ok("Flattened merged cells", f"{merged_count} region(s) → {out.name}"))
        result = {
            "success": True,
            "op": "flatten_merged_cells",
            "file": path.name,
            "sheet": sheet_name,
            "output_path": str(out),
            "merged_regions_found": merged_count,
            "rows": len(df),
            "cols": len(df.columns),
            "backup": backup,
            "progress": progress,
        }
        embed_content(result, out, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("flatten_merged_cells error")
        return {
            "success": False,
            "error": error_text(exc),
            "backup": backup,
            "hint": hint_for_error(exc, "Check the file is a valid .xlsx with merged cells."),
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }


# ---------------------------------------------------------------------------
# 10. convert_file
# ---------------------------------------------------------------------------


def convert_file(
    file_path: str,
    output_format: str = "csv",
    output_path: str = "",
    sheet: str = "",
    dry_run: bool = False,
    return_content: bool = False,
) -> dict:
    from io import BytesIO

    backup = None
    progress = []
    try:
        path = resolve_path(file_path)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {path.name}",
                "hint": "Check that file_path is absolute and the file exists.",
                "progress": [fail("File not found", str(path))],
                "token_estimate": 20,
            }
        ext = path.suffix.lower()
        if ext not in _ALL_INPUT_EXTS:
            return {
                "success": False,
                "error": f"Unsupported input format {ext!r}",
                "hint": f"Supported inputs: {sorted(_ALL_INPUT_EXTS)}",
                "progress": [fail("Unsupported input", ext)],
                "token_estimate": 20,
            }
        if output_format not in _OUTPUT_FMTS:
            return {
                "success": False,
                "error": f"Unknown output_format {output_format!r}",
                "hint": f"Valid formats: {', '.join(sorted(_OUTPUT_FMTS))}",
                "progress": [fail("Unknown output format", output_format)],
                "token_estimate": 20,
            }

        target_ext = _FMT_EXT[output_format]
        if ext == target_ext or (ext in {".xlsx", ".xls"} and output_format == "excel"):
            return {
                "success": False,
                "error": "File is already in the target format.",
                "hint": "Use normalize_headers() or trim_empty() to clean the file instead.",
                "progress": [fail("Same format", f"{ext} → {output_format}")],
                "token_estimate": 20,
            }

        # Read
        sheet_used = None
        other_sheets: list[str] = []
        if ext in {".xlsx", ".ods"}:
            pd_engine = "odf" if ext == ".ods" else "openpyxl"
            xl = pd.ExcelFile(str(path), engine=pd_engine)
            names = xl.sheet_names
            if sheet and sheet.lstrip("-").isdigit():
                idx = int(sheet)
                if idx < 0 or idx >= len(names):
                    return {
                        "success": False,
                        "error": f"Sheet index {sheet!r} out of range ({len(names)} sheet(s))",
                        "hint": f"Available: {names}. Call list_sheets() to inspect.",
                        "progress": [fail("Sheet not found", sheet)],
                        "token_estimate": 20,
                    }
                sheet_used = names[idx]
            elif sheet:
                if sheet not in names:
                    return {
                        "success": False,
                        "error": f"Sheet {sheet!r} not found",
                        "hint": f"Available: {names}. Call list_sheets() to inspect.",
                        "progress": [fail("Sheet not found", sheet)],
                        "token_estimate": 20,
                    }
                sheet_used = sheet
            else:
                sheet_used = names[0]
            other_sheets = [n for n in names if n != sheet_used]
            df = xl.parse(sheet_used)
        elif ext == ".csv":
            df = read_csv_ragged(path)
        elif ext == ".json":
            df = pd.read_json(str(path))
        elif ext == ".parquet":
            df = pd.read_parquet(str(path), engine="pyarrow")
        else:
            df = pd.DataFrame()

        out = Path(output_path) if output_path else path.parent / (path.stem + target_ext)

        if other_sheets:
            progress.append(
                warn(
                    "Other sheets ignored",
                    f"Converted only {sheet_used!r}; {len(other_sheets)} more sheet(s) not included: {other_sheets}",
                )
            )

        if dry_run:
            progress.append(info("Dry run — no changes written", path.name))
            result = {
                "success": True,
                "dry_run": True,
                "op": "convert_file",
                "would_change": {"input": path.name, "output_format": output_format, "output_path": str(out)},
                "progress": progress,
            }
            result["token_estimate"] = _token_estimate(result)
            return result

        if out.exists():
            backup = snapshot(str(out))
            progress.append(info("Snapshot created", Path(backup).name))

        out.parent.mkdir(parents=True, exist_ok=True)

        # Write
        if output_format == "csv":
            atomic_write_text(out, df.to_csv(index=False))
        elif output_format == "json":
            atomic_write_text(out, df.to_json(orient="records", indent=2))
        elif output_format == "parquet":
            buf = BytesIO()
            df.to_parquet(buf, engine="pyarrow", index=False)
            atomic_write(out, buf.getvalue())
        elif output_format == "excel":
            buf = BytesIO()
            df.to_excel(buf, index=False, engine="openpyxl")
            atomic_write(out, buf.getvalue())

        append_receipt(
            str(path),
            tool="convert_file",
            args={"output_format": output_format, "sheet": sheet_used or ""},
            result=f"converted {ext} → {target_ext} ({len(df)} rows)",
            backup=backup or "",
        )
        progress.append(ok("Converted file", f"{ext} → {target_ext}: {out.name}"))
        result = {
            "success": True,
            "op": "convert_file",
            "file": path.name,
            "input_format": ext.lstrip("."),
            "output_format": output_format,
            "output_path": str(out),
            "sheet": sheet_used,
            "other_sheets_ignored": other_sheets,
            "rows": len(df),
            "cols": len(df.columns),
            "backup": backup,
            "progress": progress,
        }
        embed_content(result, out, return_content)
        result["token_estimate"] = _token_estimate(result)
        return result
    except Exception as exc:
        logger.exception("convert_file error")
        return {
            "success": False,
            "error": error_text(exc),
            "backup": backup,
            "hint": f"Valid output formats: {', '.join(sorted(_OUTPUT_FMTS))}",
            "progress": [fail("Unexpected error", str(exc))],
            "token_estimate": 20,
        }
