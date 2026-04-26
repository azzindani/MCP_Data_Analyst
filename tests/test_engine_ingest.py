"""Tests for servers/data_ingest/engine.py."""

from __future__ import annotations

import importlib
from pathlib import Path

import pandas as pd
import pytest

from servers.data_ingest.engine import (
    convert_file,
    detect_tables,
    extract_all_sheets,
    extract_sheet,
    extract_table,
    flatten_merged_cells,
    list_sheets,
    normalize_headers,
    promote_header,
    trim_empty,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def simple_xlsx(tmp_path) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Value", "Category"])
    ws.append(["Alice", 100, "A"])
    ws.append(["Bob", 200, "B"])
    ws.append(["Carol", 150, "A"])
    ws.append(["Dave", 300, "C"])
    ws.append(["Eve", 250, "B"])
    p = tmp_path / "simple.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture()
def multi_sheet_xlsx(tmp_path) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Region", "Revenue"])
    ws1.append(["West", 5000])
    ws1.append(["East", 7000])
    ws2 = wb.create_sheet("Costs")
    ws2.append(["Region", "Cost"])
    ws2.append(["West", 2000])
    ws3 = wb.create_sheet("Meta")
    ws3.append(["Key", "Value"])
    ws3.append(["period", "Q1"])
    p = tmp_path / "multi.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture()
def multi_table_xlsx(tmp_path) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    ws.append([1, 2])
    ws.append([3, 4])
    ws.append([None, None])  # blank separator row
    ws.append(["X", "Y", "Z"])
    ws.append([10, 20, 30])
    ws.append([40, 50, 60])
    p = tmp_path / "multi_table.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture()
def merged_xlsx(tmp_path) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Region"
    ws["B1"] = "Revenue"
    ws["C1"] = "Revenue"  # will be merged with B1
    ws.merge_cells("B1:C1")
    ws.append(["West", 1000, 2000])
    ws.append(["East", 1500, 2500])
    p = tmp_path / "merged.xlsx"
    wb.save(str(p))
    return p


@pytest.fixture()
def ods_file(tmp_path) -> Path:
    df = pd.DataFrame({"Name": ["Alice", "Bob"], "Score": [95, 88]})
    p = tmp_path / "test.ods"
    df.to_excel(str(p), index=False, engine="odf")
    return p


@pytest.fixture()
def dirty_csv(tmp_path) -> Path:
    p = tmp_path / "dirty.csv"
    p.write_text(" Name ,  Revenue , Category \n,, \nAlice,100,A\nBob,200,B\n,, \n")
    return p


@pytest.fixture()
def no_header_csv(tmp_path) -> Path:
    p = tmp_path / "no_header.csv"
    p.write_text("garbage,garbage,garbage\nName,Value,Tag\nAlice,100,A\nBob,200,B\n")
    return p


@pytest.fixture()
def simple_csv(tmp_path) -> Path:
    p = tmp_path / "data.csv"
    p.write_text("Name,Value,Category\nAlice,100,A\nBob,200,B\n")
    return p


# ---------------------------------------------------------------------------
# TestListSheets
# ---------------------------------------------------------------------------


class TestListSheets:
    def test_success(self, simple_xlsx):
        r = list_sheets(str(simple_xlsx))
        assert r["success"] is True
        assert r["sheet_count"] == 1
        assert r["sheets"][0]["name"] == "Data"

    def test_multi_sheet(self, multi_sheet_xlsx):
        r = list_sheets(str(multi_sheet_xlsx))
        assert r["success"] is True
        assert r["sheet_count"] == 3
        names = [s["name"] for s in r["sheets"]]
        assert "Sales" in names and "Costs" in names and "Meta" in names

    def test_ods_file(self, ods_file):
        r = list_sheets(str(ods_file))
        assert r["success"] is True
        assert r["sheet_count"] >= 1

    def test_wrong_extension(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("a,b\n1,2\n")
        r = list_sheets(str(f))
        assert r["success"] is False
        assert "hint" in r

    def test_file_not_found(self, tmp_path):
        r = list_sheets(str(tmp_path / "missing.xlsx"))
        assert r["success"] is False
        assert "hint" in r

    def test_progress_present(self, simple_xlsx):
        r = list_sheets(str(simple_xlsx))
        assert "progress" in r
        assert len(r["progress"]) > 0

    def test_token_estimate(self, simple_xlsx):
        r = list_sheets(str(simple_xlsx))
        assert r["token_estimate"] > 0

    def test_constrained_mode(self, multi_sheet_xlsx, monkeypatch):
        monkeypatch.setenv("MCP_CONSTRAINED_MODE", "1")
        import shared.platform_utils as pu

        importlib.reload(pu)
        r = list_sheets(str(multi_sheet_xlsx))
        # 3 sheets < constrained limit of 10; should all appear
        assert r["success"] is True
        monkeypatch.delenv("MCP_CONSTRAINED_MODE", raising=False)
        importlib.reload(pu)


# ---------------------------------------------------------------------------
# TestExtractSheet
# ---------------------------------------------------------------------------


class TestExtractSheet:
    def test_success(self, simple_xlsx, tmp_path):
        out = tmp_path / "out.csv"
        r = extract_sheet(str(simple_xlsx), output_path=str(out))
        assert r["success"] is True
        assert out.exists()
        df = pd.read_csv(str(out))
        assert list(df.columns) == ["Name", "Value", "Category"]
        assert len(df) == 5

    def test_content_correct(self, simple_xlsx, tmp_path):
        out = tmp_path / "out.csv"
        extract_sheet(str(simple_xlsx), output_path=str(out))
        df = pd.read_csv(str(out))
        assert df["Name"].iloc[0] == "Alice"
        assert df["Value"].iloc[1] == 200

    def test_sheet_by_name(self, multi_sheet_xlsx, tmp_path):
        out = tmp_path / "costs.csv"
        r = extract_sheet(str(multi_sheet_xlsx), sheet="Costs", output_path=str(out))
        assert r["success"] is True
        assert r["sheet"] == "Costs"

    def test_sheet_by_index(self, multi_sheet_xlsx, tmp_path):
        out = tmp_path / "s1.csv"
        r = extract_sheet(str(multi_sheet_xlsx), sheet="1", output_path=str(out))
        assert r["success"] is True

    def test_dry_run(self, simple_xlsx, tmp_path):
        out = tmp_path / "out.csv"
        r = extract_sheet(str(simple_xlsx), output_path=str(out), dry_run=True)
        assert r["success"] is True
        assert r["dry_run"] is True
        assert "would_change" in r
        assert not out.exists()

    def test_progress_present(self, simple_xlsx, tmp_path):
        out = tmp_path / "out.csv"
        r = extract_sheet(str(simple_xlsx), output_path=str(out))
        assert "progress" in r

    def test_wrong_extension(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("a,b\n1,2\n")
        r = extract_sheet(str(f))
        assert r["success"] is False
        assert "hint" in r

    def test_file_not_found(self, tmp_path):
        r = extract_sheet(str(tmp_path / "missing.xlsx"))
        assert r["success"] is False

    def test_sheet_not_found(self, simple_xlsx):
        r = extract_sheet(str(simple_xlsx), sheet="NoSuchSheet")
        assert r["success"] is False
        assert "hint" in r

    def test_ods_extraction(self, ods_file, tmp_path):
        out = tmp_path / "out.csv"
        r = extract_sheet(str(ods_file), output_path=str(out))
        assert r["success"] is True
        df = pd.read_csv(str(out))
        assert "Name" in df.columns


# ---------------------------------------------------------------------------
# TestExtractAllSheets
# ---------------------------------------------------------------------------


class TestExtractAllSheets:
    def test_success(self, multi_sheet_xlsx, tmp_path):
        out_dir = tmp_path / "sheets"
        r = extract_all_sheets(str(multi_sheet_xlsx), output_dir=str(out_dir))
        assert r["success"] is True
        assert r["sheet_count"] == 3
        assert len(list(out_dir.glob("*.csv"))) == 3

    def test_content_correct(self, multi_sheet_xlsx, tmp_path):
        out_dir = tmp_path / "sheets"
        extract_all_sheets(str(multi_sheet_xlsx), output_dir=str(out_dir))
        csvs = list(out_dir.glob("*Sales*.csv"))
        assert len(csvs) == 1
        df = pd.read_csv(str(csvs[0]))
        assert "Revenue" in df.columns

    def test_backup_in_response(self, multi_sheet_xlsx, tmp_path):
        r = extract_all_sheets(str(multi_sheet_xlsx), output_dir=str(tmp_path / "out"))
        assert "backup" in r

    def test_dry_run(self, multi_sheet_xlsx, tmp_path):
        out_dir = tmp_path / "sheets"
        r = extract_all_sheets(str(multi_sheet_xlsx), output_dir=str(out_dir), dry_run=True)
        assert r["success"] is True
        assert r["dry_run"] is True
        assert not out_dir.exists()

    def test_progress_present(self, multi_sheet_xlsx, tmp_path):
        r = extract_all_sheets(str(multi_sheet_xlsx), output_dir=str(tmp_path / "out"))
        assert "progress" in r and len(r["progress"]) > 0

    def test_wrong_extension(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("a,b\n1,2\n")
        r = extract_all_sheets(str(f))
        assert r["success"] is False

    def test_file_not_found(self, tmp_path):
        r = extract_all_sheets(str(tmp_path / "missing.xlsx"))
        assert r["success"] is False


# ---------------------------------------------------------------------------
# TestDetectTables
# ---------------------------------------------------------------------------


class TestDetectTables:
    def test_detects_two_tables(self, multi_table_xlsx):
        r = detect_tables(str(multi_table_xlsx))
        assert r["success"] is True
        assert r["table_count"] == 2

    def test_bounding_boxes(self, multi_table_xlsx):
        r = detect_tables(str(multi_table_xlsx))
        t0 = r["tables"][0]
        assert "row_start" in t0
        assert "row_end" in t0
        assert t0["rows"] >= 2

    def test_single_table(self, simple_xlsx):
        r = detect_tables(str(simple_xlsx))
        assert r["success"] is True
        assert r["table_count"] == 1

    def test_min_rows_filter(self, multi_table_xlsx):
        r = detect_tables(str(multi_table_xlsx), min_rows=5)
        # Table 1 has 3 rows, table 2 has 3 rows; neither meets min_rows=5
        assert r["success"] is True
        assert r["table_count"] == 0

    def test_wrong_extension(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("a,b\n1,2\n")
        r = detect_tables(str(f))
        assert r["success"] is False

    def test_file_not_found(self, tmp_path):
        r = detect_tables(str(tmp_path / "missing.xlsx"))
        assert r["success"] is False

    def test_sheet_not_found(self, simple_xlsx):
        r = detect_tables(str(simple_xlsx), sheet="NoSheet")
        assert r["success"] is False
        assert "hint" in r

    def test_progress_present(self, simple_xlsx):
        r = detect_tables(str(simple_xlsx))
        assert "progress" in r

    def test_token_estimate(self, simple_xlsx):
        r = detect_tables(str(simple_xlsx))
        assert r["token_estimate"] > 0


# ---------------------------------------------------------------------------
# TestExtractTable
# ---------------------------------------------------------------------------


class TestExtractTable:
    def test_success(self, multi_table_xlsx, tmp_path):
        out = tmp_path / "table0.csv"
        r = extract_table(str(multi_table_xlsx), table_index=0, output_path=str(out))
        assert r["success"] is True
        assert out.exists()

    def test_second_table(self, multi_table_xlsx, tmp_path):
        out = tmp_path / "table1.csv"
        r = extract_table(str(multi_table_xlsx), table_index=1, output_path=str(out))
        assert r["success"] is True
        assert out.exists()

    def test_dry_run(self, multi_table_xlsx, tmp_path):
        out = tmp_path / "table0.csv"
        r = extract_table(str(multi_table_xlsx), table_index=0, output_path=str(out), dry_run=True)
        assert r["success"] is True
        assert r["dry_run"] is True
        assert not out.exists()

    def test_index_out_of_range(self, multi_table_xlsx):
        r = extract_table(str(multi_table_xlsx), table_index=99)
        assert r["success"] is False
        assert "hint" in r

    def test_backup_present_on_overwrite(self, multi_table_xlsx, tmp_path):
        out = tmp_path / "table0.csv"
        # First extract creates the file
        extract_table(str(multi_table_xlsx), table_index=0, output_path=str(out))
        # Second extract should snapshot the existing file
        r = extract_table(str(multi_table_xlsx), table_index=0, output_path=str(out))
        assert r["success"] is True
        assert r["backup"] is not None
        assert ".mcp_versions" in r["backup"]

    def test_progress_present(self, multi_table_xlsx, tmp_path):
        r = extract_table(str(multi_table_xlsx), table_index=0)
        assert "progress" in r

    def test_wrong_extension(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("a,b\n1,2\n")
        r = extract_table(str(f))
        assert r["success"] is False

    def test_file_not_found(self, tmp_path):
        r = extract_table(str(tmp_path / "missing.xlsx"))
        assert r["success"] is False


# ---------------------------------------------------------------------------
# TestNormalizeHeaders
# ---------------------------------------------------------------------------


class TestNormalizeHeaders:
    def test_success(self, dirty_csv):
        r = normalize_headers(str(dirty_csv))
        assert r["success"] is True
        assert r["renamed_count"] > 0

    def test_content_correct(self, dirty_csv):
        normalize_headers(str(dirty_csv))
        df = pd.read_csv(str(dirty_csv))
        assert "name" in df.columns
        assert "revenue" in df.columns

    def test_snapshot_created(self, dirty_csv):
        r = normalize_headers(str(dirty_csv))
        assert ".mcp_versions" in r["backup"]
        versions_dir = dirty_csv.parent / ".mcp_versions"
        assert versions_dir.exists()

    def test_backup_in_response(self, dirty_csv):
        r = normalize_headers(str(dirty_csv))
        assert "backup" in r

    def test_dry_run(self, dirty_csv):
        original = dirty_csv.read_text()
        r = normalize_headers(str(dirty_csv), dry_run=True)
        assert r["success"] is True
        assert r["dry_run"] is True
        assert "would_change" in r
        assert dirty_csv.read_text() == original

    def test_progress_present(self, dirty_csv):
        r = normalize_headers(str(dirty_csv))
        assert "progress" in r and len(r["progress"]) > 0

    def test_wrong_extension(self, simple_xlsx):
        r = normalize_headers(str(simple_xlsx))
        assert r["success"] is False
        assert "hint" in r

    def test_file_not_found(self, tmp_path):
        r = normalize_headers(str(tmp_path / "missing.csv"))
        assert r["success"] is False

    def test_deduplication(self, tmp_path):
        f = tmp_path / "dup.csv"
        f.write_text("Value,Value,Value\n1,2,3\n")
        r = normalize_headers(str(f))
        assert r["success"] is True
        df = pd.read_csv(str(f))
        assert len(set(df.columns)) == 3  # all unique after dedup

    def test_token_estimate(self, dirty_csv):
        r = normalize_headers(str(dirty_csv))
        assert r["token_estimate"] > 0


# ---------------------------------------------------------------------------
# TestTrimEmpty
# ---------------------------------------------------------------------------


class TestTrimEmpty:
    def test_success(self, dirty_csv):
        r = trim_empty(str(dirty_csv))
        assert r["success"] is True

    def test_rows_dropped(self, dirty_csv):
        r = trim_empty(str(dirty_csv))
        assert r["rows_dropped"] > 0

    def test_content_correct(self, dirty_csv):
        trim_empty(str(dirty_csv))
        df = pd.read_csv(str(dirty_csv))
        # All remaining rows should have at least one non-empty value
        assert len(df) > 0

    def test_snapshot_created(self, dirty_csv):
        r = trim_empty(str(dirty_csv))
        assert ".mcp_versions" in r["backup"]

    def test_backup_in_response(self, dirty_csv):
        r = trim_empty(str(dirty_csv))
        assert "backup" in r

    def test_dry_run(self, dirty_csv):
        original = dirty_csv.read_text()
        r = trim_empty(str(dirty_csv), dry_run=True)
        assert r["success"] is True
        assert r["dry_run"] is True
        assert dirty_csv.read_text() == original

    def test_progress_present(self, dirty_csv):
        r = trim_empty(str(dirty_csv))
        assert "progress" in r

    def test_wrong_extension(self, simple_xlsx):
        r = trim_empty(str(simple_xlsx))
        assert r["success"] is False
        assert "hint" in r

    def test_file_not_found(self, tmp_path):
        r = trim_empty(str(tmp_path / "missing.csv"))
        assert r["success"] is False

    def test_token_estimate(self, dirty_csv):
        r = trim_empty(str(dirty_csv))
        assert r["token_estimate"] > 0


# ---------------------------------------------------------------------------
# TestPromoteHeader
# ---------------------------------------------------------------------------


class TestPromoteHeader:
    def test_success(self, no_header_csv):
        r = promote_header(str(no_header_csv), row_index=1)
        assert r["success"] is True

    def test_content_correct(self, no_header_csv):
        promote_header(str(no_header_csv), row_index=1)
        df = pd.read_csv(str(no_header_csv))
        assert "Name" in df.columns
        assert "Value" in df.columns

    def test_snapshot_created(self, no_header_csv):
        r = promote_header(str(no_header_csv), row_index=1)
        assert ".mcp_versions" in r["backup"]

    def test_backup_in_response(self, no_header_csv):
        r = promote_header(str(no_header_csv), row_index=1)
        assert "backup" in r

    def test_dry_run(self, no_header_csv):
        original = no_header_csv.read_text()
        r = promote_header(str(no_header_csv), row_index=1, dry_run=True)
        assert r["success"] is True
        assert r["dry_run"] is True
        assert "would_change" in r
        assert no_header_csv.read_text() == original

    def test_progress_present(self, no_header_csv):
        r = promote_header(str(no_header_csv), row_index=1)
        assert "progress" in r

    def test_wrong_extension(self, simple_xlsx):
        r = promote_header(str(simple_xlsx))
        assert r["success"] is False
        assert "hint" in r

    def test_file_not_found(self, tmp_path):
        r = promote_header(str(tmp_path / "missing.csv"))
        assert r["success"] is False

    def test_row_out_of_range(self, simple_csv):
        r = promote_header(str(simple_csv), row_index=999)
        assert r["success"] is False
        assert "hint" in r

    def test_token_estimate(self, no_header_csv):
        r = promote_header(str(no_header_csv), row_index=1)
        assert r["token_estimate"] > 0


# ---------------------------------------------------------------------------
# TestFlattenMergedCells
# ---------------------------------------------------------------------------


class TestFlattenMergedCells:
    def test_success(self, merged_xlsx, tmp_path):
        out = tmp_path / "flat.csv"
        r = flatten_merged_cells(str(merged_xlsx), output_path=str(out))
        assert r["success"] is True
        assert out.exists()

    def test_merged_regions_detected(self, merged_xlsx, tmp_path):
        out = tmp_path / "flat.csv"
        r = flatten_merged_cells(str(merged_xlsx), output_path=str(out))
        assert r["merged_regions_found"] >= 1

    def test_content_correct(self, merged_xlsx, tmp_path):
        out = tmp_path / "flat.csv"
        flatten_merged_cells(str(merged_xlsx), output_path=str(out))
        df = pd.read_csv(str(out))
        # B1 and C1 both had "Revenue"; after flattening both cols should have same header-derived name
        assert len(df.columns) >= 2

    def test_backup_present_on_overwrite(self, merged_xlsx, tmp_path):
        out = tmp_path / "flat.csv"
        flatten_merged_cells(str(merged_xlsx), output_path=str(out))
        r = flatten_merged_cells(str(merged_xlsx), output_path=str(out))
        assert r["success"] is True
        assert r["backup"] is not None

    def test_dry_run(self, merged_xlsx, tmp_path):
        out = tmp_path / "flat.csv"
        r = flatten_merged_cells(str(merged_xlsx), output_path=str(out), dry_run=True)
        assert r["success"] is True
        assert r["dry_run"] is True
        assert not out.exists()

    def test_progress_present(self, merged_xlsx, tmp_path):
        r = flatten_merged_cells(str(merged_xlsx), output_path=str(tmp_path / "flat.csv"))
        assert "progress" in r

    def test_wrong_extension(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("a,b\n1,2\n")
        r = flatten_merged_cells(str(f))
        assert r["success"] is False
        assert "hint" in r

    def test_file_not_found(self, tmp_path):
        r = flatten_merged_cells(str(tmp_path / "missing.xlsx"))
        assert r["success"] is False

    def test_sheet_not_found(self, merged_xlsx):
        r = flatten_merged_cells(str(merged_xlsx), sheet="NoSheet")
        assert r["success"] is False
        assert "hint" in r


# ---------------------------------------------------------------------------
# TestConvertFile
# ---------------------------------------------------------------------------


class TestConvertFile:
    def test_xlsx_to_csv(self, simple_xlsx, tmp_path):
        out = tmp_path / "out.csv"
        r = convert_file(str(simple_xlsx), output_format="csv", output_path=str(out))
        assert r["success"] is True
        assert out.exists()
        df = pd.read_csv(str(out))
        assert list(df.columns) == ["Name", "Value", "Category"]

    def test_csv_to_json(self, simple_csv, tmp_path):
        out = tmp_path / "out.json"
        r = convert_file(str(simple_csv), output_format="json", output_path=str(out))
        assert r["success"] is True
        assert out.exists()

    def test_csv_to_parquet(self, simple_csv, tmp_path):
        out = tmp_path / "out.parquet"
        r = convert_file(str(simple_csv), output_format="parquet", output_path=str(out))
        assert r["success"] is True
        assert out.exists()
        df = pd.read_parquet(str(out))
        assert "Name" in df.columns

    def test_xlsx_to_json(self, simple_xlsx, tmp_path):
        out = tmp_path / "out.json"
        r = convert_file(str(simple_xlsx), output_format="json", output_path=str(out))
        assert r["success"] is True

    def test_ods_to_csv(self, ods_file, tmp_path):
        out = tmp_path / "out.csv"
        r = convert_file(str(ods_file), output_format="csv", output_path=str(out))
        assert r["success"] is True
        assert out.exists()

    def test_backup_on_overwrite(self, simple_xlsx, tmp_path):
        out = tmp_path / "out.csv"
        convert_file(str(simple_xlsx), output_format="csv", output_path=str(out))
        r = convert_file(str(simple_xlsx), output_format="csv", output_path=str(out))
        assert r["success"] is True
        assert r["backup"] is not None
        assert ".mcp_versions" in r["backup"]

    def test_dry_run(self, simple_xlsx, tmp_path):
        out = tmp_path / "out.csv"
        r = convert_file(str(simple_xlsx), output_format="csv", output_path=str(out), dry_run=True)
        assert r["success"] is True
        assert r["dry_run"] is True
        assert not out.exists()

    def test_progress_present(self, simple_xlsx, tmp_path):
        r = convert_file(str(simple_xlsx), output_format="csv", output_path=str(tmp_path / "out.csv"))
        assert "progress" in r

    def test_unknown_output_format(self, simple_csv):
        r = convert_file(str(simple_csv), output_format="xml")
        assert r["success"] is False
        assert "hint" in r

    def test_file_not_found(self, tmp_path):
        r = convert_file(str(tmp_path / "missing.xlsx"), output_format="csv")
        assert r["success"] is False

    def test_same_format_rejected(self, simple_csv):
        r = convert_file(str(simple_csv), output_format="csv")
        assert r["success"] is False
        assert "hint" in r

    def test_token_estimate(self, simple_xlsx, tmp_path):
        r = convert_file(str(simple_xlsx), output_format="csv", output_path=str(tmp_path / "out.csv"))
        assert r["token_estimate"] > 0
